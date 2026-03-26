from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook

from app.schemas.email_delivery import EmailSendResult
from app.services.email_service import EmailService
from app.services.official_company_registry_service import OfficialCompanyRegistryService
from app.services.errors import ProviderRequestError
from app.services.twilio_voice_service import TwilioCallDispatchResult, TwilioVoiceService


pytestmark = pytest.mark.anyio


def build_valid_record(external_id: str = "1", email: str | None = None) -> dict:
    payload = {
        "external_id": external_id,
        "client_name": "Empresa Exemplo LTDA",
        "cnpj": "11.222.333/0001-81",
        "phone": "11987654321",
    }
    if email is not None:
        payload["email"] = email
    return payload


async def test_create_validation_batch_returns_received_snapshot_for_rails_polling(client):
    payload = {
        "id_lote": "lote_001",
        "origem": "web",
        "records": [
            {
                "id_registro": "1",
                "nome_cliente": "Empresa Exemplo LTDA",
                "cnpj": "11.222.333/0001-81",
                "telefone": "(11) 98765-4321",
            },
            {
                "external_id": "2",
                "client_name": "Cliente Telefone Invalido",
                "cnpj": "11.222.333/0001-81",
                "phone": "1234",
            },
        ],
    }

    response = await client.post("/validations", json=payload)

    assert response.status_code == 202

    data = response.json()
    assert data["batch_id"] == "lote_001"
    assert data["source"] == "web"
    assert data["batch_status"] == "processing"
    assert data["result_ready"] is False
    assert data["technical_status"] == "processing"
    assert data["finished_at"] is None
    assert data["total_records"] == 2
    assert data["summary"] == {
        "ready_for_call": 1,
        "ready_for_retry_call": 0,
        "validation_failed": 1,
        "invalid_phone": 1,
        "cnpj_not_found": 0,
        "processing": 1,
        "pending_records": 1,
        "validated_records": 0,
        "failed_records": 1,
        "confirmed_by_call": 0,
        "confirmed_by_whatsapp": 0,
        "waiting_whatsapp_reply": 0,
        "confirmed_by_email": 0,
        "waiting_email_reply": 0,
    }

    first_record = data["records"][0]
    assert first_record["external_id"] == "1"
    assert first_record["cnpj_normalized"] == "11222333000181"
    assert first_record["phone_normalized"] == "5511987654321"
    assert first_record["call_status"] == "not_started"
    assert first_record["call_result"] == "not_started"
    assert first_record["final_status"] == "processing"
    assert first_record["official_registry_email"] == "contato@empresaexemplo.com.br"
    assert first_record["email_status"] == "not_required"
    assert first_record["call_attempts"] == []
    assert first_record["whatsapp_history"] == []
    assert first_record["email_history"] == []

    second_record = data["records"][1]
    assert second_record["phone_valid"] is False
    assert second_record["business_status"] == "invalid_phone"
    assert second_record["final_status"] == "validation_failed"


async def test_dispatch_moves_batch_to_processing_and_creates_initial_call_attempt(client):
    payload = {
        "batch_id": "lote_dispatch",
        "source": "web",
        "records": [build_valid_record()],
    }

    await client.post("/validations", json=payload)

    dispatch_response = await client.post("/validations/lote_dispatch/dispatch")

    assert dispatch_response.status_code == 200
    data = dispatch_response.json()
    assert data["batch_status"] == "processing"
    assert data["result_ready"] is False
    assert data["summary"]["pending_records"] == 1
    assert data["records"][0]["call_status"] == "queued"
    assert data["records"][0]["call_result"] == "pending_dispatch"
    assert len(data["records"][0]["call_attempts"]) == 1
    assert data["records"][0]["call_attempts"][0]["status"] == "queued"
    assert data["records"][0]["call_attempts"][0]["phone_source"] == "payload_phone"


async def test_stop_test_batch_cancels_pending_attempts_and_blocks_new_dispatches(
    client,
    monkeypatch,
):
    dispatched_calls: list[str] = []
    ended_calls: list[str] = []

    monkeypatch.setattr(TwilioVoiceService, "is_configured", lambda self: True)

    def fake_create_outbound_call(self, **kwargs) -> TwilioCallDispatchResult:
        provider_call_id = f"CA-stop-{kwargs['external_id']}-{kwargs['attempt_number']}"
        dispatched_calls.append(provider_call_id)
        assert kwargs["caller_company_name"] == "Central de Validacao Cadastral"
        return TwilioCallDispatchResult(
            provider_call_id=provider_call_id,
            provider_status="queued",
            raw_payload={"sid": provider_call_id, "status": "queued"},
        )

    def fake_end_outbound_call(self, *, provider_call_id: str) -> None:
        ended_calls.append(provider_call_id)

    monkeypatch.setattr(TwilioVoiceService, "create_outbound_call", fake_create_outbound_call)
    monkeypatch.setattr(TwilioVoiceService, "end_outbound_call", fake_end_outbound_call)

    payload = {
        "batch_id": "lote_stop_test_ui",
        "source": "web",
        "records": [build_valid_record("1"), build_valid_record("2")],
    }

    await client.post("/validations", json=payload)
    await client.post("/validations/lote_stop_test_ui/dispatch")

    stop_response = await client.post("/test/voice-call/batch/lote_stop_test_ui/stop")

    assert stop_response.status_code == 200
    data = stop_response.json()
    assert ended_calls == ["CA-stop-1-1"]
    assert dispatched_calls == ["CA-stop-1-1"]
    assert data["batch_status"] == "processing"
    assert data["records"][0]["final_status"] == "processing"
    assert data["records"][1]["final_status"] == "validation_failed"
    assert data["records"][1]["call_status"] == "failed"
    assert data["records"][1]["call_result"] == "inconclusive"
    assert data["records"][1]["call_attempts"][0]["status"] == "failed"
    assert data["records"][1]["call_attempts"][0]["result"] == "inconclusive"

    final_event_response = await client.post(
        "/validations/lote_stop_test_ui/records/1/call-events",
        json={
            "provider_call_id": "CA-stop-1-1",
            "call_status": "failed",
            "call_result": "inconclusive",
            "transcript_summary": "Ligacao encerrada manualmente no teste.",
            "duration_seconds": 4,
        },
    )

    assert final_event_response.status_code == 200
    final_data = final_event_response.json()
    assert final_data["final_status"] == "validation_failed"
    assert dispatched_calls == ["CA-stop-1-1"]

    batch_result = await client.get("/validations/lote_stop_test_ui")
    batch_data = batch_result.json()
    assert batch_data["batch_status"] == "completed"
    assert batch_data["summary"]["failed_records"] == 2


async def test_dispatch_uses_twilio_when_provider_is_configured(client, monkeypatch):
    monkeypatch.setattr(TwilioVoiceService, "is_configured", lambda self: True)

    def fake_create_outbound_call(self, **kwargs) -> TwilioCallDispatchResult:
        assert kwargs["batch_id"] == "lote_twilio"
        assert kwargs["external_id"] == "1"
        assert kwargs["attempt_number"] == 1
        assert kwargs["client_name"] == "Empresa Exemplo LTDA"
        assert kwargs["cnpj"] == "11222333000181"
        assert kwargs["phone_to_dial"] == "5511987654321"
        assert kwargs["caller_company_name"] == "Central de Validacao Cadastral"
        assert kwargs["twiml_mode"] == "media_stream"
        return TwilioCallDispatchResult(
            provider_call_id="CA1234567890",
            provider_status="queued",
            raw_payload={"sid": "CA1234567890", "status": "queued"},
        )

    monkeypatch.setattr(TwilioVoiceService, "create_outbound_call", fake_create_outbound_call)

    payload = {
        "batch_id": "lote_twilio",
        "source": "web",
        "records": [build_valid_record()],
    }

    await client.post("/validations", json=payload)
    dispatch_response = await client.post("/validations/lote_twilio/dispatch")

    assert dispatch_response.status_code == 200
    data = dispatch_response.json()
    assert data["records"][0]["call_attempts"][0]["provider_call_id"] == "CA1234567890"
    assert "Twilio Voice" not in data["records"][0]["observation"]


async def test_rejected_call_queues_retry_on_alternative_phone_found_in_registry(
    client,
    monkeypatch,
):
    def fake_fetch_company_data(self, cnpj: str | None):
        assert cnpj == "11222333000181"
        return {
            "cnpj": "11222333000181",
            "razao_social": "EMPRESA EXEMPLO LTDA",
            "ddd_telefone_1": "11999990000",
            "ddd_telefone_2": "",
        }

    monkeypatch.setattr(
        OfficialCompanyRegistryService,
        "fetch_company_data",
        fake_fetch_company_data,
    )

    payload = {
        "batch_id": "lote_retry_phone",
        "source": "web",
        "records": [build_valid_record()],
    }

    await client.post("/validations", json=payload)
    await client.post("/validations/lote_retry_phone/dispatch")
    batch_response = await client.get("/validations/lote_retry_phone")
    provider_call_id = batch_response.json()["records"][0]["call_attempts"][0]["provider_call_id"]

    response = await client.post(
        "/validations/lote_retry_phone/records/1/call-events",
        json={
            "provider_call_id": provider_call_id,
            "call_status": "answered",
            "call_result": "rejected",
            "transcript_summary": "Contato informou que o numero nao pertence a empresa.",
            "sentiment": "neutral",
            "duration_seconds": 35,
        },
    )

    assert response.status_code == 200
    data = response.json()
    assert data["business_status"] == "ready_for_retry_call"
    assert data["call_status"] == "queued"
    assert data["call_result"] == "pending_dispatch"
    assert data["final_status"] == "processing"
    assert data["whatsapp_status"] == "not_required"
    assert len(data["call_attempts"]) == 2
    assert data["call_attempts"][1]["phone_dialed"] == "5511999990000"
    assert data["call_attempts"][1]["phone_source"] == "official_company_registry"
    assert data["official_registry_checked"] is True
    assert data["official_registry_retry_found"] is True
    assert data["official_registry_retry_phone"] == "5511999990000"
    assert data["last_phone_dialed"] == "5511999990000"
    assert data["last_phone_source"] == "official_company_registry"
    assert data["attempted_phones"] == ["5511987654321", "5511999990000"]

    batch_result = await client.get("/validations/lote_retry_phone")
    batch_data = batch_result.json()
    assert batch_data["batch_status"] == "processing"
    assert batch_data["summary"]["ready_for_retry_call"] == 1
    assert batch_data["summary"]["waiting_whatsapp_reply"] == 0


async def test_official_registry_retry_phone_is_not_retried_when_twilio_trial_rejects_unverified_number(
    client,
    monkeypatch,
):
    monkeypatch.setattr(TwilioVoiceService, "is_configured", lambda self: True)

    def fake_fetch_company_data(self, cnpj: str | None):
        return {
            "cnpj": "11222333000181",
            "razao_social": "EMPRESA EXEMPLO LTDA",
            "ddd_telefone_1": "5136354333",
            "ddd_telefone_2": "",
        }

    dispatched_phones: list[str] = []

    def fake_create_outbound_call(self, **kwargs):
        dispatched_phones.append(str(kwargs["phone_to_dial"]))
        if str(kwargs["phone_to_dial"]) == "555136354333":
            raise ProviderRequestError(
                "Twilio Voice",
                "Twilio retornou erro ao criar a ligacao. status_code=400 response_body={\"code\":21219}",
                status_code=400,
                provider_code="21219",
            )
        return TwilioCallDispatchResult(
            provider_call_id=f"CA-{kwargs['attempt_number']}",
            provider_status="queued",
            raw_payload={"sid": f"CA-{kwargs['attempt_number']}", "status": "queued"},
        )

    monkeypatch.setattr(OfficialCompanyRegistryService, "fetch_company_data", fake_fetch_company_data)
    monkeypatch.setattr(TwilioVoiceService, "create_outbound_call", fake_create_outbound_call)

    payload = {
        "batch_id": "lote_trial_retry_unverified",
        "source": "web",
        "records": [build_valid_record()],
    }

    await client.post("/validations", json=payload)
    await client.post("/validations/lote_trial_retry_unverified/dispatch")
    batch_response = await client.get("/validations/lote_trial_retry_unverified")
    provider_call_id = batch_response.json()["records"][0]["call_attempts"][0]["provider_call_id"]

    response = await client.post(
        "/validations/lote_trial_retry_unverified/records/1/call-events",
        json={
            "provider_call_id": provider_call_id,
            "call_status": "answered",
            "call_result": "rejected",
            "transcript_summary": "Contato informou que o numero nao pertence a empresa.",
            "sentiment": "neutral",
            "duration_seconds": 20,
        },
    )

    assert response.status_code == 200
    data = response.json()
    assert dispatched_phones == ["5511987654321", "555136354333"]
    assert data["final_status"] == "validation_failed"
    assert data["business_status"] == "rejected_by_call"
    assert data["call_attempts"][1]["status"] == "failed"
    assert data["call_attempts"][1]["result"] == "inconclusive"
    assert "trial do Twilio" in data["observation"]

    batch_result = await client.get("/validations/lote_trial_retry_unverified")
    batch_data = batch_result.json()
    assert batch_data["batch_status"] == "completed"
    assert batch_data["summary"]["failed_records"] == 1
    assert batch_data["summary"]["ready_for_retry_call"] == 0


async def test_twilio_completed_after_answer_keeps_record_processing_until_media_stream_result(
    client,
):
    payload = {
        "batch_id": "lote_completed_after_answer",
        "source": "web",
        "records": [build_valid_record()],
    }

    await client.post("/validations", json=payload)
    await client.post("/validations/lote_completed_after_answer/dispatch")
    batch_response = await client.get("/validations/lote_completed_after_answer")
    provider_call_id = batch_response.json()["records"][0]["call_attempts"][0]["provider_call_id"]

    answered_response = await client.post(
        "/webhooks/twilio/voice/status?batch_id=lote_completed_after_answer&external_id=1&attempt_number=1",
        data={
            "CallSid": provider_call_id,
            "CallStatus": "in-progress",
            "CallDuration": "12",
        },
    )

    assert answered_response.status_code == 200

    completed_response = await client.post(
        "/webhooks/twilio/voice/status?batch_id=lote_completed_after_answer&external_id=1&attempt_number=1",
        data={
            "CallSid": provider_call_id,
            "CallStatus": "completed",
            "CallDuration": "24",
        },
    )

    assert completed_response.status_code == 200

    intermediate_batch_response = await client.get("/validations/lote_completed_after_answer")
    data = intermediate_batch_response.json()
    assert data["batch_status"] == "processing"
    assert data["result_ready"] is False
    assert data["summary"]["pending_records"] == 1
    assert data["records"][0]["final_status"] == "processing"
    assert data["records"][0]["call_status"] == "answered"
    assert data["records"][0]["business_status"] == "call_answered"

    final_response = await client.post(
        "/validations/lote_completed_after_answer/records/1/call-events",
        json={
            "provider_call_id": provider_call_id,
            "call_status": "answered",
            "call_result": "confirmed",
            "transcript_summary": "cliente: sim continua sendo da empresa | agente: obrigada validacao concluida com sucesso",
            "duration_seconds": 24,
            "observation": "Numero confirmado por ligacao conversacional.",
        },
    )

    assert final_response.status_code == 200
    final_data = final_response.json()
    assert final_data["final_status"] == "validated"
    assert final_data["business_status"] == "confirmed_by_call"
    assert final_data["phone_confirmed"] is True
    assert final_data["validated_phone"] == "5511987654321"
    assert final_data["customer_transcript"] == "sim continua sendo da empresa"
    assert final_data["assistant_transcript"] == "obrigada validacao concluida com sucesso"
    assert final_data["attempted_phones"] == ["5511987654321"]
    assert final_data["call_attempts"][0]["customer_transcript"] == "sim continua sendo da empresa"
    assert final_data["call_attempts"][0]["assistant_transcript"] == "obrigada validacao concluida com sucesso"


async def test_twilio_completed_without_media_stream_finalizes_batch_with_inconclusive_error(
    client,
):
    payload = {
        "batch_id": "lote_completed_without_media",
        "source": "web",
        "records": [build_valid_record()],
    }

    await client.post("/validations", json=payload)
    await client.post("/validations/lote_completed_without_media/dispatch")
    batch_response = await client.get("/validations/lote_completed_without_media")
    provider_call_id = batch_response.json()["records"][0]["call_attempts"][0]["provider_call_id"]

    response = await client.post(
        "/webhooks/twilio/voice/status?batch_id=lote_completed_without_media&external_id=1&attempt_number=1",
        data={
            "CallSid": provider_call_id,
            "CallStatus": "completed",
            "CallDuration": "24",
        },
    )

    assert response.status_code == 200

    final_batch_response = await client.get("/validations/lote_completed_without_media")
    data = final_batch_response.json()
    assert data["batch_status"] == "completed"
    assert data["result_ready"] is True
    assert data["summary"]["pending_records"] == 0
    assert data["summary"]["failed_records"] == 1
    assert data["records"][0]["final_status"] == "validation_failed"
    assert data["records"][0]["business_status"] == "inconclusive_call"
    assert data["records"][0]["call_result"] == "inconclusive"


async def test_batch_is_completed_immediately_when_all_records_fail_initial_validation(client):
    payload = {
        "batch_id": "lote_invalidos",
        "source": "web",
        "records": [
            {
                "external_id": "1",
                "client_name": "Cliente Invalido",
                "cnpj": "123",
                "phone": "999",
            }
        ],
    }

    response = await client.post("/validations", json=payload)

    assert response.status_code == 202
    data = response.json()
    assert data["batch_status"] == "completed"
    assert data["result_ready"] is True
    assert data["finished_at"] is not None
    assert data["summary"]["pending_records"] == 0
    assert data["summary"]["failed_records"] == 1
    assert data["records"][0]["final_status"] == "validation_failed"


async def test_create_validation_batch_returns_payload_error_for_invalid_body(client):
    payload = {"batch_id": "lote_002", "source": "web", "records": []}

    response = await client.post("/validations", json=payload)

    assert response.status_code == 422

    data = response.json()
    assert data["message"] == "Payload inválido."
    assert data["technical_status"] == "payload_invalid"
    assert data["errors"]


async def test_create_validation_batch_rejects_duplicate_batch_id(client):
    payload = {
        "batch_id": "lote_duplicado",
        "source": "web",
        "records": [build_valid_record()],
    }

    first_response = await client.post("/validations", json=payload)
    second_response = await client.post("/validations", json=payload)

    assert first_response.status_code == 202
    assert second_response.status_code == 409
    assert second_response.json() == {"detail": "Lote 'lote_duplicado' ja existe."}


async def test_get_validation_batch_returns_404_when_batch_does_not_exist(client):
    response = await client.get("/validations/lote_inexistente")

    assert response.status_code == 404
    assert response.json() == {"detail": "Lote 'lote_inexistente' nao encontrado."}


async def test_call_event_can_confirm_record_and_complete_batch(client):
    payload = {
        "batch_id": "lote_call_confirmed",
        "source": "web",
        "records": [build_valid_record()],
    }

    await client.post("/validations", json=payload)
    await client.post("/validations/lote_call_confirmed/dispatch")
    batch_response = await client.get("/validations/lote_call_confirmed")
    provider_call_id = batch_response.json()["records"][0]["call_attempts"][0]["provider_call_id"]

    response = await client.post(
        "/validations/lote_call_confirmed/records/1/call-events",
        json={
            "provider_call_id": provider_call_id,
            "call_status": "answered",
            "call_result": "confirmed",
            "transcript_summary": "Contato confirmou que o numero pertence a empresa.",
            "sentiment": "neutral",
            "duration_seconds": 42,
        },
    )

    assert response.status_code == 200
    data = response.json()
    assert data["business_status"] == "confirmed_by_call"
    assert data["final_status"] == "validated"
    assert data["confirmation_source"] == "voice_call"
    assert data["phone_confirmed"] is True
    assert data["call_attempts"][0]["status"] == "answered"
    assert data["call_attempts"][0]["result"] == "confirmed"

    batch_result = await client.get("/validations/lote_call_confirmed")
    batch_data = batch_result.json()
    assert batch_data["batch_status"] == "completed"


async def test_twilio_status_callback_marks_record_as_not_answered_and_sends_email_fallback(
    client,
    monkeypatch,
):
    monkeypatch.setattr(TwilioVoiceService, "is_configured", lambda self: True)
    monkeypatch.setattr(
        TwilioVoiceService,
        "create_outbound_call",
        lambda self, **kwargs: TwilioCallDispatchResult(
            provider_call_id="CA999",
            provider_status="queued",
            raw_payload={"sid": "CA999", "status": "queued"},
        ),
    )
    monkeypatch.setattr(
        EmailService,
        "send_validation_fallback_email",
        lambda self, **kwargs: EmailSendResult(
            success=True,
            provider_message_id="email_test_1",
            subject="Validacao cadastral da empresa Empresa Exemplo LTDA",
            message_body="Fallback por e-mail de teste.",
        ),
    )

    payload = {
        "batch_id": "lote_no_answer",
        "source": "web",
        "records": [build_valid_record()],
    }

    await client.post("/validations", json=payload)
    await client.post("/validations/lote_no_answer/dispatch")

    response = await client.post(
        "/webhooks/twilio/voice/status?batch_id=lote_no_answer&external_id=1&attempt_number=1",
        data={
            "CallSid": "CA999",
            "CallStatus": "no-answer",
            "CallDuration": "0",
        },
    )

    assert response.status_code == 200
    batch_response = await client.get("/validations/lote_no_answer")
    data = batch_response.json()
    assert data["records"][0]["business_status"] == "waiting_email_reply"
    assert data["records"][0]["email_status"] == "waiting_email_reply"
    assert data["records"][0]["fallback_email_used"] == "contato@empresaexemplo.com.br"
    assert len(data["records"][0]["email_history"]) == 1
    assert data["records"][0]["email_history"][0]["provider_message_id"] == "email_test_1"
    assert data["records"][0]["whatsapp_history"] == []
    assert data["summary"]["waiting_email_reply"] == 1


async def test_test_voice_call_start_creates_batch_and_dispatches(client, monkeypatch):
    monkeypatch.setattr(TwilioVoiceService, "is_configured", lambda self: True)
    monkeypatch.setattr(
        TwilioVoiceService,
        "create_outbound_call",
        lambda self, **kwargs: TwilioCallDispatchResult(
            provider_call_id="CATESTCALL",
            provider_status="queued",
            raw_payload={"sid": "CATESTCALL", "status": "queued"},
        ),
    )

    response = await client.post(
        "/test/voice-call/start",
        json={
            "client_name": "Empresa Exemplo LTDA",
            "cnpj": "11.222.333/0001-81",
            "phone": "11987654321",
            "call_scenario": "failed",
            "fallback_message": "Mensagem de teste",
        },
    )

    assert response.status_code == 200
    data = response.json()
    assert data["batch_id"].startswith("test_voice_")
    assert data["records"][0]["call_attempts"][0]["provider_call_id"] == "CATESTCALL"


async def test_test_voice_call_start_accepts_realtime_profile_overrides(client, monkeypatch):
    monkeypatch.setattr(TwilioVoiceService, "is_configured", lambda self: True)
    captured_kwargs: dict[str, object] = {}

    def fake_create_outbound_call(self, **kwargs) -> TwilioCallDispatchResult:
        captured_kwargs.update(kwargs)
        return TwilioCallDispatchResult(
            provider_call_id="CATESTPROFILE",
            provider_status="queued",
            raw_payload={"sid": "CATESTPROFILE", "status": "queued"},
        )

    monkeypatch.setattr(TwilioVoiceService, "create_outbound_call", fake_create_outbound_call)

    response = await client.post(
        "/test/voice-call/start?realtime_model=gpt-realtime-1.5&realtime_voice=cedar&realtime_output_speed=0.93&realtime_style_profile=warm_feminine",
        json={
            "client_name": "Empresa Exemplo LTDA",
            "cnpj": "11.222.333/0001-81",
            "phone": "11987654321",
            "call_scenario": "failed",
            "fallback_message": "Mensagem de teste",
        },
    )

    assert response.status_code == 200
    assert captured_kwargs["realtime_model_override"] == "gpt-realtime-1.5"
    assert captured_kwargs["realtime_voice_override"] == "cedar"
    assert captured_kwargs["realtime_output_speed_override"] == 0.93
    assert captured_kwargs["realtime_style_profile"] == "warm_feminine"


async def test_multi_record_dispatch_only_starts_first_call_in_queue(client, monkeypatch):
    monkeypatch.setattr(TwilioVoiceService, "is_configured", lambda self: True)
    dispatched_calls: list[str] = []

    def fake_create_outbound_call(self, **kwargs) -> TwilioCallDispatchResult:
        dispatched_calls.append(kwargs["external_id"])
        return TwilioCallDispatchResult(
            provider_call_id=f"CA_{kwargs['external_id']}",
            provider_status="queued",
            raw_payload={"sid": f"CA_{kwargs['external_id']}", "status": "queued"},
        )

    monkeypatch.setattr(TwilioVoiceService, "create_outbound_call", fake_create_outbound_call)

    payload = {
        "batch_id": "lote_fila_inicial",
        "source": "web",
        "records": [build_valid_record("1"), build_valid_record("2")],
    }

    await client.post("/validations", json=payload)
    response = await client.post("/validations/lote_fila_inicial/dispatch")

    assert response.status_code == 200
    assert dispatched_calls == ["1"]

    data = response.json()
    assert len(data["records"][0]["call_attempts"]) == 1
    assert len(data["records"][1]["call_attempts"]) == 1
    assert data["records"][0]["call_attempts"][0]["provider_call_id"] == "CA_1"
    assert data["records"][1]["call_attempts"][0]["provider_call_id"].startswith("call_")
    assert data["records"][1]["call_attempts"][0]["result"] == "pending_dispatch"


async def test_finishing_first_record_dispatches_next_call_from_batch_queue(client, monkeypatch):
    monkeypatch.setattr(TwilioVoiceService, "is_configured", lambda self: True)
    dispatched_calls: list[str] = []

    def fake_create_outbound_call(self, **kwargs) -> TwilioCallDispatchResult:
        dispatched_calls.append(kwargs["external_id"])
        return TwilioCallDispatchResult(
            provider_call_id=f"CA_{kwargs['external_id']}_{kwargs['attempt_number']}",
            provider_status="queued",
            raw_payload={"sid": f"CA_{kwargs['external_id']}_{kwargs['attempt_number']}", "status": "queued"},
        )

    monkeypatch.setattr(TwilioVoiceService, "create_outbound_call", fake_create_outbound_call)

    payload = {
        "batch_id": "lote_fila_sequencial",
        "source": "web",
        "records": [build_valid_record("1"), build_valid_record("2")],
    }

    await client.post("/validations", json=payload)
    await client.post("/validations/lote_fila_sequencial/dispatch")

    response = await client.post(
        "/validations/lote_fila_sequencial/records/1/call-events",
        json={
            "provider_call_id": "CA_1_1",
            "call_status": "answered",
            "call_result": "confirmed",
            "transcript_summary": "Contato confirmou que o numero pertence a empresa.",
            "duration_seconds": 18,
        },
    )

    assert response.status_code == 200
    assert dispatched_calls == ["1", "2"]

    batch_response = await client.get("/validations/lote_fila_sequencial")
    batch_data = batch_response.json()
    assert batch_data["records"][0]["final_status"] == "validated"
    assert batch_data["records"][1]["call_attempts"][0]["provider_call_id"] == "CA_2_1"
    assert batch_data["records"][1]["call_status"] == "queued"


async def test_test_batch_voice_call_start_imports_xlsx_and_dispatches_first_record(client, monkeypatch):
    monkeypatch.setattr(TwilioVoiceService, "is_configured", lambda self: True)
    monkeypatch.setattr(
        TwilioVoiceService,
        "create_outbound_call",
        lambda self, **kwargs: TwilioCallDispatchResult(
            provider_call_id=f"CA_BATCH_{kwargs['external_id']}",
            provider_status="queued",
            raw_payload={"sid": f"CA_BATCH_{kwargs['external_id']}", "status": "queued"},
        ),
    )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["nome_cliente", "cnpj", "telefone"])
    worksheet.append(["Empresa Exemplo LTDA", "11.222.333/0001-81", "11987654321"])
    worksheet.append(["Empresa Exemplo LTDA", "11.222.333/0001-81", "11987654321"])

    file_buffer = BytesIO()
    workbook.save(file_buffer)
    file_buffer.seek(0)

    response = await client.post(
        "/test/voice-call/batch/start",
        files={
            "file": (
                "lote_teste.xlsx",
                file_buffer.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    data = response.json()
    assert data["batch_id"].startswith("test_voice_batch_")
    assert data["total_records"] == 2
    assert data["records"][0]["call_attempts"][0]["provider_call_id"] == "CA_BATCH_2"
    assert data["records"][1]["call_attempts"][0]["provider_call_id"].startswith("call_")


async def test_test_batch_voice_call_start_accepts_distinct_structurally_valid_cnpjs_in_homologation_mode(client, monkeypatch):
    monkeypatch.setattr(TwilioVoiceService, "is_configured", lambda self: True)
    monkeypatch.setattr(
        TwilioVoiceService,
        "create_outbound_call",
        lambda self, **kwargs: TwilioCallDispatchResult(
            provider_call_id=f"CA_HOMOLOG_{kwargs['external_id']}",
            provider_status="queued",
            raw_payload={"sid": f"CA_HOMOLOG_{kwargs['external_id']}", "status": "queued"},
        ),
    )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["external_id", "client_name", "cnpj", "phone"])
    worksheet.append(["1", "Empresa Alfa LTDA", "12.345.678/0001-95", "11987654321"])
    worksheet.append(["2", "Empresa Beta LTDA", "23.456.789/0001-95", "11987654321"])

    file_buffer = BytesIO()
    workbook.save(file_buffer)
    file_buffer.seek(0)

    response = await client.post(
        "/test/voice-call/batch/start?skip_registry_validation=true",
        files={
            "file": (
                "lote_homologacao.xlsx",
                file_buffer.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    data = response.json()
    assert data["total_records"] == 2
    assert data["records"][0]["ready_for_contact"] is True
    assert data["records"][1]["ready_for_contact"] is True
    assert data["records"][0]["call_attempts"][0]["provider_call_id"] == "CA_HOMOLOG_1"
    assert data["records"][1]["call_attempts"][0]["provider_call_id"].startswith("call_")


async def test_test_batch_results_spreadsheet_download_returns_current_batch_snapshot(client):
    payload = {
        "batch_id": "lote_download_planilha_teste",
        "source": "web",
        "records": [build_valid_record()],
    }

    await client.post("/validations", json=payload)
    await client.post("/validations/lote_download_planilha_teste/dispatch")
    batch_response = await client.get("/validations/lote_download_planilha_teste")
    provider_call_id = batch_response.json()["records"][0]["call_attempts"][0]["provider_call_id"]

    await client.post(
        "/validations/lote_download_planilha_teste/records/1/call-events",
        json={
            "provider_call_id": provider_call_id,
            "call_status": "answered",
            "call_result": "confirmed",
            "transcript_summary": "cliente: Pertence sim. | agente: Obrigada pela confirmacao. Validacao concluida com sucesso.",
            "duration_seconds": 21,
        },
    )

    response = await client.get("/test/voice-call/batch/lote_download_planilha_teste/results.xlsx")

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert (
        'attachment; filename="lote_download_planilha_teste_resultado_validacao.xlsx"'
        in response.headers["content-disposition"]
    )

    workbook = load_workbook(BytesIO(response.content))
    assert workbook.sheetnames == ["Retorno", "Resumo", "Tentativas", "Emails"]

    return_sheet = workbook["Retorno"]
    headers = [cell.value for cell in return_sheet[1]]
    row_values = [cell.value for cell in return_sheet[2]]
    row = dict(zip(headers, row_values))
    assert row["lote_id"] == "lote_download_planilha_teste"
    assert row["id_registro"] == "1"
    assert row["empresa"] == "Empresa Exemplo LTDA"
    assert row["cnpj_informado"] == "11.222.333/0001-81"
    assert row["telefone_informado"] == "11987654321"
    assert row["telefone_normalizado"] == "5511987654321"
    assert row["email_base_oficial"] == "contato@empresaexemplo.com.br"
    assert row["status_email"] == "not_required"
    assert row["telefone_validado"] == "5511987654321"
    assert row["telefone_confirmado"] == "True"
    assert row["origem_confirmacao"] == "voice_call"
    assert row["consultou_base_oficial"] == "False"
    assert row["transcricao_cliente"] == "Pertence sim."
    assert row["transcricao_agente"] == "Obrigada pela confirmacao. Validacao concluida com sucesso."
    assert row["status_final"] == "validated"

    summary_sheet = workbook["Resumo"]
    assert summary_sheet["A2"].value == "batch_id"
    assert summary_sheet["B2"].value == "lote_download_planilha_teste"

    attempts_sheet = workbook["Tentativas"]
    attempt_headers = [cell.value for cell in attempts_sheet[1]]
    attempt_values = [cell.value for cell in attempts_sheet[2]]
    attempt_row = dict(zip(attempt_headers, attempt_values))
    assert attempt_row["phone_dialed"] == "5511987654321"
    assert attempt_row["customer_transcript"] == "Pertence sim."
    assert attempt_row["assistant_transcript"] == "Obrigada pela confirmacao. Validacao concluida com sucesso."

    emails_sheet = workbook["Emails"]
    email_headers = [cell.value for cell in emails_sheet[1]]
    assert email_headers == [
        "external_id",
        "provider_message_id",
        "recipient_email",
        "direction",
        "status",
        "subject",
        "sent_at",
        "responded_at",
        "response_text",
        "observation",
    ]
    assert emails_sheet.max_row == 1
