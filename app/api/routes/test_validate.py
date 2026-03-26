from datetime import datetime, timezone
import csv
import io
import logging
from typing import Any, Literal
from uuid import uuid4

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse

from openpyxl import Workbook, load_workbook

from ..dependencies import (
    get_local_test_flow_service,
    get_supplier_discovery_service,
    get_validation_async_service,
    get_validation_flow_service,
)
from ...schemas.request import SupplierValidationBatchRequest, ValidationBatchRequest
from ...schemas.response import ValidationBatchResponse
from ...schemas.test_flow import (
    ClearStateResponse,
    LocalTestFlowResponse,
    LocalTestStateResponse,
    LocalValidationRequest,
    SupplierDiscoveryRequest,
    SupplierDiscoveryResponse,
    ManualWhatsAppSendRequest,
    WhatsAppSendResult,
)
from ...services.errors import BatchNotFoundError
from ...services.local_test_flow_service import LocalTestFlowService
from ...services.supplier_discovery_service import SupplierDiscoveryService
from ...services.validation_async_service import ValidationAsyncService
from ...services.validation_flow import ValidationFlowService
from ...services.validation_snapshot_builder import ValidationSnapshotBuilder

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/test", tags=["local-test-flow"])


class _BypassOfficialCompanyRegistryService:
    def exists(self, cnpj: str | None) -> bool:
        logger.info(
            "Modo de homologacao do lote ignorando consulta inicial da base oficial | cnpj=%s",
            cnpj,
        )
        return True

    def find_contact_email(self, *, cnpj: str | None) -> None:
        logger.info(
            "Modo de homologacao do lote ignorando busca inicial de e-mail na base oficial | cnpj=%s",
            cnpj,
        )
        return None


def _create_batch_for_test_ui(
    *,
    batch_request: ValidationBatchRequest,
    validation_flow_service: ValidationFlowService,
    skip_registry_validation: bool,
    caller_company_name: str | None = None,
) -> ValidationBatchResponse:
    if not skip_registry_validation:
        return validation_flow_service.create_batch(
            batch_request,
            caller_company_name=caller_company_name,
        )

    test_snapshot_builder = ValidationSnapshotBuilder(
        official_company_registry=_BypassOfficialCompanyRegistryService(),
    )
    test_validation_flow_service = ValidationFlowService(
        snapshot_builder=test_snapshot_builder,
        batch_repository=validation_flow_service.batch_repository,
    )
    logger.warning(
        "Modo de homologacao do lote habilitado: consulta inicial da base oficial sera ignorada | batch_id=%s",
        batch_request.batch_id,
    )
    return test_validation_flow_service.create_batch(
        batch_request,
        caller_company_name=caller_company_name,
    )


_BATCH_HEADER_ALIASES = {
    "external_id": {"external_id", "id_registro", "id", "registro"},
    "client_name": {
        "client_name",
        "supplier_name",
        "nome_cliente",
        "nome_fornecedor",
        "empresa",
        "cliente",
        "razao_social",
        "nome",
    },
    "cnpj": {"cnpj"},
    "phone": {"phone", "telefone", "celular", "fone", "whatsapp"},
    "email": {"email", "e_mail", "correio_eletronico", "mail"},
}




_SUPPLIER_BATCH_HEADER_ALIASES = {
    "external_id": {"external_id", "id_registro", "id", "registro"},
    "supplier_name": {
        "supplier_name",
        "client_name",
        "nome_fornecedor",
        "nome_cliente",
        "empresa",
        "fornecedor",
        "nome",
    },
    "phone": {"phone", "telefone", "celular", "fone", "whatsapp"},
    "email": {"email", "e_mail", "correio_eletronico", "mail"},
}

def _normalize_header_name(value: Any) -> str:
    return "_".join(str(value or "").strip().lower().replace("-", " ").split())


def _normalize_cell_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _map_headers(headers: list[str]) -> dict[str, int]:
    mapped_headers: dict[str, int] = {}

    for index, header in enumerate(headers):
        normalized_header = _normalize_header_name(header)
        for field_name, aliases in _BATCH_HEADER_ALIASES.items():
            if normalized_header in aliases and field_name not in mapped_headers:
                mapped_headers[field_name] = index
                break

    missing_required_headers = [
        field_name
        for field_name in ("client_name", "cnpj", "phone")
        if field_name not in mapped_headers
    ]
    if missing_required_headers:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                "Planilha sem colunas obrigatorias. Encontradas: "
                f"{headers}. Obrigatorias: client_name/nome_cliente, cnpj, phone/telefone."
            ),
        )

    return mapped_headers


def _build_record_from_values(
    row_values: list[str],
    header_positions: dict[str, int],
    row_number: int,
) -> dict[str, str] | None:
    if not any(value for value in row_values):
        return None

    def read(field_name: str) -> str:
        index = header_positions.get(field_name)
        if index is None or index >= len(row_values):
            return ""
        return row_values[index]

    client_name = read("client_name")
    cnpj = read("cnpj")
    phone = read("phone")
    email = read("email")
    external_id = read("external_id") or str(row_number)

    if not client_name and not cnpj and not phone and not email:
        return None

    return {
        "external_id": external_id,
        "client_name": client_name,
        "cnpj": cnpj,
        "phone": phone,
        "email": email,
    }


def _parse_csv_records(file_bytes: bytes) -> list[dict[str, str]]:
    text_content = file_bytes.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text_content))
    rows = list(reader)
    if not rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Arquivo CSV vazio.",
        )

    headers = [_normalize_cell_value(value) for value in rows[0]]
    header_positions = _map_headers(headers)
    records: list[dict[str, str]] = []
    for row_number, row in enumerate(rows[1:], start=2):
        row_values = [_normalize_cell_value(value) for value in row]
        record = _build_record_from_values(row_values, header_positions, row_number)
        if record is not None:
            records.append(record)

    return records


def _parse_xlsx_records(file_bytes: bytes) -> list[dict[str, str]]:
    workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    if not rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Arquivo XLSX vazio.",
        )

    headers = [_normalize_cell_value(value) for value in rows[0]]
    header_positions = _map_headers(headers)
    records: list[dict[str, str]] = []
    for row_number, row in enumerate(rows[1:], start=2):
        row_values = [_normalize_cell_value(value) for value in row]
        record = _build_record_from_values(row_values, header_positions, row_number)
        if record is not None:
            records.append(record)

    return records




def _map_supplier_headers(headers: list[str]) -> dict[str, int]:
    mapped_headers: dict[str, int] = {}

    for index, header in enumerate(headers):
        normalized_header = _normalize_header_name(header)
        for field_name, aliases in _SUPPLIER_BATCH_HEADER_ALIASES.items():
            if normalized_header in aliases and field_name not in mapped_headers:
                mapped_headers[field_name] = index
                break

    missing_required_headers = [
        field_name
        for field_name in ("supplier_name", "phone")
        if field_name not in mapped_headers
    ]
    if missing_required_headers:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                "Planilha de fornecedor sem colunas obrigatorias. Encontradas: "
                f"{headers}. Obrigatorias: supplier_name/nome_fornecedor e phone/telefone."
            ),
        )

    return mapped_headers


def _build_supplier_record_from_values(
    row_values: list[str],
    header_positions: dict[str, int],
    row_number: int,
) -> dict[str, str] | None:
    if not any(value for value in row_values):
        return None

    def read(field_name: str) -> str:
        index = header_positions.get(field_name)
        if index is None or index >= len(row_values):
            return ""
        return row_values[index]

    supplier_name = read("supplier_name")
    phone = read("phone")
    email = read("email")
    external_id = read("external_id") or str(row_number)

    if not supplier_name and not phone and not email:
        return None

    return {
        "external_id": external_id,
        "supplier_name": supplier_name,
        "phone": phone,
        "email": email,
    }


def _parse_supplier_csv_records(file_bytes: bytes) -> list[dict[str, str]]:
    text_content = file_bytes.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text_content))
    rows = list(reader)
    if not rows:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Arquivo CSV vazio.")

    headers = [_normalize_cell_value(value) for value in rows[0]]
    header_positions = _map_supplier_headers(headers)
    records: list[dict[str, str]] = []
    for row_number, row in enumerate(rows[1:], start=2):
        row_values = [_normalize_cell_value(value) for value in row]
        record = _build_supplier_record_from_values(row_values, header_positions, row_number)
        if record is not None:
            records.append(record)

    return records


def _parse_supplier_xlsx_records(file_bytes: bytes) -> list[dict[str, str]]:
    workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    if not rows:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Arquivo XLSX vazio.")

    headers = [_normalize_cell_value(value) for value in rows[0]]
    header_positions = _map_supplier_headers(headers)
    records: list[dict[str, str]] = []
    for row_number, row in enumerate(rows[1:], start=2):
        row_values = [_normalize_cell_value(value) for value in row]
        record = _build_supplier_record_from_values(row_values, header_positions, row_number)
        if record is not None:
            records.append(record)

    return records


def _parse_supplier_records_from_upload(filename: str, file_bytes: bytes) -> list[dict[str, str]]:
    lowered_name = filename.lower()
    if lowered_name.endswith(".csv"):
        records = _parse_supplier_csv_records(file_bytes)
    elif lowered_name.endswith(".xlsx"):
        records = _parse_supplier_xlsx_records(file_bytes)
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Formato nao suportado. Envie um arquivo .xlsx ou .csv.",
        )

    if not records:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Nenhum registro valido foi encontrado na planilha de fornecedor.",
        )

    return records


def _normalize_optional_text(value: str | None) -> str | None:
    if value is None:
        return None
    cleaned_value = value.strip()
    return cleaned_value or None


def _normalize_optional_speed(value: float | None) -> float | None:
    return value


def _format_sheet_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value)


def _build_batch_result_workbook(batch_snapshot: ValidationBatchResponse) -> bytes:
    workbook = Workbook()

    return_sheet = workbook.active
    return_sheet.title = "Retorno"
    return_sheet.append([
        "lote_id",
        "id_registro",
        "empresa",
        "cnpj_informado",
        "cnpj_normalizado",
        "telefone_informado",
        "telefone_normalizado",
        "email_informado",
        "email_normalizado",
        "email_base_oficial",
        "email_fallback_utilizado",
        "status_email",
        "workflow_kind",
        "segmento_fornecedor",
        "fornecedor_telefone_pertence_empresa",
        "fornecedor_fornece_segmento",
        "fornecedor_interesse_comercial",
        "fornecedor_callback_phone",
        "fornecedor_outcome",
        "telefone_validado",
        "ultimo_telefone_discado",
        "origem_ultimo_telefone",
        "tipo_telefone",
        "cnpj_encontrado_base",
        "telefone_estruturalmente_valido",
        "pronto_para_contato",
        "status_tecnico",
        "status_negocio",
        "status_ligacao",
        "resultado_ligacao",
        "status_whatsapp",
        "telefone_confirmado",
        "origem_confirmacao",
        "consultou_base_oficial",
        "encontrou_telefone_alternativo_base",
        "telefone_alternativo_base",
        "quantidade_tentativas",
        "telefones_tentados",
        "transcricao_cliente",
        "transcricao_agente",
        "transcricao_resumo",
        "status_final",
        "observacao",
    ])
    for record in batch_snapshot.records:
        return_sheet.append([
            batch_snapshot.batch_id,
            record.external_id,
            record.client_name,
            record.cnpj_original,
            record.cnpj_normalized,
            record.phone_original,
            record.phone_normalized,
            record.email_original,
            record.email_normalized,
            record.official_registry_email,
            record.fallback_email_used,
            record.email_status,
            batch_snapshot.workflow_kind,
            (record.supplier_validation.segment_name if record.supplier_validation else batch_snapshot.segment_name),
            _format_sheet_value(record.supplier_validation.phone_belongs_to_company if record.supplier_validation else None),
            _format_sheet_value(record.supplier_validation.supplies_segment if record.supplier_validation else None),
            _format_sheet_value(record.supplier_validation.commercial_interest if record.supplier_validation else None),
            (record.supplier_validation.callback_phone_informed if record.supplier_validation else batch_snapshot.callback_phone),
            (record.supplier_validation.outcome if record.supplier_validation else None),
            record.validated_phone,
            record.last_phone_dialed,
            record.last_phone_source,
            record.phone_type,
            _format_sheet_value(record.cnpj_found),
            _format_sheet_value(record.phone_valid),
            _format_sheet_value(record.ready_for_contact),
            record.technical_status,
            record.business_status,
            record.call_status,
            record.call_result,
            record.whatsapp_status,
            _format_sheet_value(record.phone_confirmed),
            record.confirmation_source,
            _format_sheet_value(record.official_registry_checked),
            _format_sheet_value(record.official_registry_retry_found),
            record.official_registry_retry_phone,
            record.attempts_count,
            ", ".join(record.attempted_phones),
            record.customer_transcript,
            record.assistant_transcript,
            record.transcript_summary,
            record.final_status,
            record.observation,
        ])

    summary_sheet = workbook.create_sheet(title="Resumo")
    summary_sheet.append(["campo", "valor"])
    summary_sheet.append(["batch_id", batch_snapshot.batch_id])
    summary_sheet.append(["source", batch_snapshot.source])
    summary_sheet.append(["batch_status", batch_snapshot.batch_status])
    summary_sheet.append(["technical_status", batch_snapshot.technical_status])
    summary_sheet.append(["created_at", _format_sheet_value(batch_snapshot.created_at)])
    summary_sheet.append(["updated_at", _format_sheet_value(batch_snapshot.updated_at)])
    summary_sheet.append(["finished_at", _format_sheet_value(batch_snapshot.finished_at)])
    summary_sheet.append(["result_ready", _format_sheet_value(batch_snapshot.result_ready)])
    summary_sheet.append(["total_records", _format_sheet_value(batch_snapshot.total_records)])
    summary_sheet.append([])
    summary_sheet.append(["metrica", "valor"])
    for key, value in batch_snapshot.summary.model_dump().items():
        summary_sheet.append([key, value])

    attempts_sheet = workbook.create_sheet(title="Tentativas")
    attempts_sheet.append([
        "external_id",
        "attempt_number",
        "provider_call_id",
        "phone_dialed",
        "phone_source",
        "status",
        "result",
        "started_at",
        "finished_at",
        "duration_seconds",
        "customer_transcript",
        "assistant_transcript",
        "transcript_summary",
        "observation",
    ])
    for record in batch_snapshot.records:
        for attempt in record.call_attempts:
            attempts_sheet.append([
                record.external_id,
                attempt.attempt_number,
                attempt.provider_call_id,
                attempt.phone_dialed,
                attempt.phone_source,
                attempt.status,
                attempt.result,
                _format_sheet_value(attempt.started_at),
                _format_sheet_value(attempt.finished_at),
                attempt.duration_seconds,
                attempt.customer_transcript,
                attempt.assistant_transcript,
                attempt.transcript_summary,
                attempt.observation,
            ])

    emails_sheet = workbook.create_sheet(title="Emails")
    emails_sheet.append([
        "external_id",
        "provider_message_id",
        "recipient_email",
        "direction",
        "status",
        "subject",
        "sent_at",
        "responded_at",
        "response_text",
        "observation",
    ])
    for record in batch_snapshot.records:
        for email_message in record.email_history:
            emails_sheet.append([
                record.external_id,
                email_message.provider_message_id,
                email_message.recipient_email,
                email_message.direction,
                email_message.status,
                email_message.subject,
                _format_sheet_value(email_message.sent_at),
                _format_sheet_value(email_message.responded_at),
                email_message.response_text,
                email_message.observation,
            ])

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()



def _build_supplier_discovery_workbook(search_result: SupplierDiscoveryResponse) -> bytes:
    workbook = Workbook()

    results_sheet = workbook.active
    results_sheet.title = 'Fornecedores'
    results_sheet.append([
        'search_id',
        'segment_name',
        'region',
        'supplier_name',
        'phone',
        'website',
        'city',
        'state',
        'source_urls',
        'discovery_confidence',
        'notes',
        'callback_phone',
        'callback_contact_name',
    ])

    for supplier in search_result.suppliers:
        results_sheet.append([
            search_result.search_id,
            search_result.segment_name,
            search_result.region,
            supplier.supplier_name,
            supplier.phone,
            supplier.website,
            supplier.city,
            supplier.state,
            ' | '.join(supplier.source_urls),
            supplier.discovery_confidence,
            supplier.notes,
            search_result.callback_phone,
            search_result.callback_contact_name,
        ])

    summary_sheet = workbook.create_sheet(title='Resumo')
    summary_sheet.append(['campo', 'valor'])
    summary_sheet.append(['search_id', search_result.search_id])
    summary_sheet.append(['mode', search_result.mode])
    summary_sheet.append(['segment_name', search_result.segment_name])
    summary_sheet.append(['region', search_result.region])
    summary_sheet.append(['generated_at', _format_sheet_value(search_result.generated_at)])
    summary_sheet.append(['total_suppliers', search_result.total_suppliers])
    summary_sheet.append(['callback_phone', search_result.callback_phone])
    summary_sheet.append(['callback_contact_name', search_result.callback_contact_name])
    summary_sheet.append(['downloadable_file_url', search_result.downloadable_file_url])
    summary_sheet.append(['message', search_result.message])

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()

def _parse_batch_records_from_upload(filename: str, file_bytes: bytes) -> list[dict[str, str]]:
    lowered_name = filename.lower()
    if lowered_name.endswith(".csv"):
        records = _parse_csv_records(file_bytes)
    elif lowered_name.endswith(".xlsx"):
        records = _parse_xlsx_records(file_bytes)
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Formato nao suportado. Envie um arquivo .xlsx ou .csv.",
        )

    if not records:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Nenhum registro valido foi encontrado na planilha.",
        )

    return records




@router.post('/supplier-discovery/search', response_model=SupplierDiscoveryResponse)
async def start_supplier_discovery_search(
    payload: SupplierDiscoveryRequest,
    service: SupplierDiscoveryService = Depends(get_supplier_discovery_service),
) -> SupplierDiscoveryResponse:
    logger.info(
        'HTTP POST /test/supplier-discovery/search recebido | segment_name=%s region=%s max_suppliers=%s',
        payload.segment_name,
        payload.region,
        payload.max_suppliers,
    )
    return await service.discover_suppliers(payload)


@router.get('/supplier-discovery/{search_id}', response_model=SupplierDiscoveryResponse)
async def get_supplier_discovery_search(
    search_id: str,
    service: SupplierDiscoveryService = Depends(get_supplier_discovery_service),
) -> SupplierDiscoveryResponse:
    logger.info(
        'HTTP GET /test/supplier-discovery/{search_id} recebido | search_id=%s',
        search_id,
    )
    search_result = service.get_discovery_result(search_id)
    if search_result is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail='Busca de fornecedores nao encontrada.',
        )
    return search_result


@router.get('/supplier-discovery/{search_id}/results.xlsx')
async def download_supplier_discovery_results(
    search_id: str,
    service: SupplierDiscoveryService = Depends(get_supplier_discovery_service),
) -> StreamingResponse:
    logger.info(
        'HTTP GET /test/supplier-discovery/{search_id}/results.xlsx recebido | search_id=%s',
        search_id,
    )
    search_result = service.get_discovery_result(search_id)
    if search_result is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail='Busca de fornecedores nao encontrada.',
        )

    workbook_bytes = _build_supplier_discovery_workbook(search_result)
    filename = f'{search_id}_fornecedores_encontrados.xlsx'
    return StreamingResponse(
        io.BytesIO(workbook_bytes),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': f'attachment; filename="{filename}"',
        },
    )

@router.post("/validate", response_model=LocalTestFlowResponse)
async def simulate_validation(
    payload: LocalValidationRequest,
    service: LocalTestFlowService = Depends(get_local_test_flow_service),
) -> LocalTestFlowResponse:
    logger.info(
        "HTTP POST /test/validate recebido | client_name=%s call_scenario=%s",
        payload.client_name,
        payload.call_scenario,
    )
    return await service.simulate_validation(payload)


@router.post("/voice-call/start", response_model=ValidationBatchResponse)
async def start_real_voice_call(
    payload: LocalValidationRequest,
    twiml_mode: Literal["media_stream", "diagnostic_say"] = Query(
        default="media_stream",
    ),
    realtime_model: str | None = Query(default=None),
    realtime_voice: str | None = Query(default=None),
    realtime_output_speed: float | None = Query(default=None),
    realtime_style_profile: str | None = Query(default=None),
    caller_company_name: str | None = Query(default=None),
    validation_flow_service: ValidationFlowService = Depends(get_validation_flow_service),
    validation_async_service: ValidationAsyncService = Depends(get_validation_async_service),
) -> ValidationBatchResponse:
    batch_id = (
        f"test_voice_{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}_{uuid4().hex[:6]}"
    )
    logger.info(
        "HTTP POST /test/voice-call/start recebido | batch_id=%s client_name=%s phone=%s twiml_mode=%s",
        batch_id,
        payload.client_name,
        payload.phone,
        twiml_mode,
    )
    batch_request = ValidationBatchRequest.model_validate(
        {
            "batch_id": batch_id,
            "source": "web",
            "records": [
                {
                    "external_id": "1",
                    "client_name": payload.client_name,
                    "cnpj": payload.cnpj,
                    "phone": payload.phone,
                }
            ],
        }
    )
    validation_flow_service.create_batch(batch_request)
    return validation_async_service.dispatch_batch(
        batch_id,
        twiml_mode=twiml_mode,
        realtime_model_override=_normalize_optional_text(realtime_model),
        realtime_voice_override=_normalize_optional_text(realtime_voice),
        realtime_output_speed_override=_normalize_optional_speed(realtime_output_speed),
        realtime_style_profile=_normalize_optional_text(realtime_style_profile),
    )


@router.post("/voice-call/batch/start", response_model=ValidationBatchResponse)
async def start_real_voice_call_batch(
    file: UploadFile = File(...),
    twiml_mode: Literal["media_stream", "diagnostic_say"] = Query(
        default="media_stream",
    ),
    skip_registry_validation: bool = Query(default=True),
    realtime_model: str | None = Query(default=None),
    realtime_voice: str | None = Query(default=None),
    realtime_output_speed: float | None = Query(default=None),
    realtime_style_profile: str | None = Query(default=None),
    caller_company_name: str | None = Query(default=None),
    validation_flow_service: ValidationFlowService = Depends(get_validation_flow_service),
    validation_async_service: ValidationAsyncService = Depends(get_validation_async_service),
) -> ValidationBatchResponse:
    batch_id = (
        f"test_voice_batch_{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}_{uuid4().hex[:6]}"
    )
    file_bytes = await file.read()
    logger.info(
        "HTTP POST /test/voice-call/batch/start recebido | batch_id=%s filename=%s twiml_mode=%s skip_registry_validation=%s",
        batch_id,
        file.filename,
        twiml_mode,
        skip_registry_validation,
    )
    records = _parse_batch_records_from_upload(file.filename or "", file_bytes)
    batch_request = ValidationBatchRequest.model_validate(
        {
            "batch_id": batch_id,
            "source": "web",
            "records": records,
        }
    )
    _create_batch_for_test_ui(
        batch_request=batch_request,
        validation_flow_service=validation_flow_service,
        skip_registry_validation=skip_registry_validation,
        caller_company_name=_normalize_optional_text(caller_company_name),
    )
    return validation_async_service.dispatch_batch(
        batch_id,
        twiml_mode=twiml_mode,
        realtime_model_override=_normalize_optional_text(realtime_model),
        realtime_voice_override=_normalize_optional_text(realtime_voice),
        realtime_output_speed_override=_normalize_optional_speed(realtime_output_speed),
        realtime_style_profile=_normalize_optional_text(realtime_style_profile),
    )




@router.post("/voice-call/supplier-batch/start", response_model=ValidationBatchResponse)
async def start_real_supplier_voice_call_batch(
    file: UploadFile = File(...),
    segment_name: str = Query(..., min_length=1),
    callback_phone: str = Query(..., min_length=1),
    callback_contact_name: str | None = Query(default=None),
    twiml_mode: Literal["media_stream", "diagnostic_say"] = Query(default="media_stream"),
    caller_company_name: str | None = Query(default=None),
    realtime_model: str | None = Query(default=None),
    realtime_voice: str | None = Query(default=None),
    realtime_output_speed: float | None = Query(default=None),
    realtime_style_profile: str | None = Query(default=None),
    validation_flow_service: ValidationFlowService = Depends(get_validation_flow_service),
    validation_async_service: ValidationAsyncService = Depends(get_validation_async_service),
) -> ValidationBatchResponse:
    batch_id = f"test_supplier_batch_{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}_{uuid4().hex[:6]}"
    file_bytes = await file.read()
    logger.info(
        "HTTP POST /test/voice-call/supplier-batch/start recebido | batch_id=%s filename=%s twiml_mode=%s segment_name=%s",
        batch_id,
        file.filename,
        twiml_mode,
        segment_name,
    )
    records = _parse_supplier_records_from_upload(file.filename or "", file_bytes)
    batch_request = SupplierValidationBatchRequest.model_validate(
        {
            "batch_id": batch_id,
            "source": "web",
            "segment_name": segment_name,
            "callback_phone": callback_phone,
            "callback_contact_name": callback_contact_name,
            "records": records,
        }
    )
    validation_flow_service.create_supplier_batch(
        batch_request,
        caller_company_name=_normalize_optional_text(caller_company_name),
    )
    return validation_async_service.dispatch_batch(
        batch_id,
        twiml_mode=twiml_mode,
        realtime_model_override=_normalize_optional_text(realtime_model),
        realtime_voice_override=_normalize_optional_text(realtime_voice),
        realtime_output_speed_override=_normalize_optional_speed(realtime_output_speed),
        realtime_style_profile=_normalize_optional_text(realtime_style_profile),
    )


@router.post("/voice-call/batch/{batch_id}/stop", response_model=ValidationBatchResponse)
async def stop_real_voice_call_batch(
    batch_id: str,
    validation_async_service: ValidationAsyncService = Depends(get_validation_async_service),
) -> ValidationBatchResponse:
    logger.warning(
        "HTTP POST /test/voice-call/batch/{batch_id}/stop recebido | batch_id=%s",
        batch_id,
    )
    return validation_async_service.stop_batch(batch_id)




@router.get("/voice-call/batch/{batch_id}/results.xlsx")
async def download_real_voice_call_batch_results(
    batch_id: str,
    validation_flow_service: ValidationFlowService = Depends(get_validation_flow_service),
) -> StreamingResponse:
    logger.info(
        "HTTP GET /test/voice-call/batch/{batch_id}/results.xlsx recebido | batch_id=%s",
        batch_id,
    )
    try:
        batch_snapshot = validation_flow_service.get_batch(batch_id)
    except BatchNotFoundError as error:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=str(error),
        ) from error

    workbook_bytes = _build_batch_result_workbook(batch_snapshot)
    filename = f"{batch_id}_resultado_validacao.xlsx"
    return StreamingResponse(
        io.BytesIO(workbook_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
        },
    )

@router.post("/whatsapp/send", response_model=WhatsAppSendResult)
async def send_whatsapp_message(
    payload: ManualWhatsAppSendRequest,
    service: LocalTestFlowService = Depends(get_local_test_flow_service),
) -> WhatsAppSendResult:
    logger.info("HTTP POST /test/whatsapp/send recebido | phone=%s", payload.phone)
    return await service.send_manual_whatsapp(payload)


@router.get("/state", response_model=LocalTestStateResponse)
async def get_local_test_state(
    service: LocalTestFlowService = Depends(get_local_test_flow_service),
) -> LocalTestStateResponse:
    return service.get_state()


@router.post("/logs/clear", response_model=ClearStateResponse)
async def clear_local_test_state(
    service: LocalTestFlowService = Depends(get_local_test_flow_service),
) -> ClearStateResponse:
    logger.info("HTTP POST /test/logs/clear recebido")
    return service.clear_state()
